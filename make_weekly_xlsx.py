import datetime
import glob
import os

import pandas as pd
import openpyxl

basedata_folder_path = 'H:/Youtube/basedata/'
src = glob.glob(basedata_folder_path + '*.xlsx')[-1]
print(src)
management = 'H:/Youtube/管理.xlsx'

# 期間の取得、ついでに期間を書き込む
wb_management = openpyxl.load_workbook(management)
sheet_m = wb_management['Sheet1']
last = sheet_m.max_row
span_start = sheet_m.cell(row=last, column=2).value
span_end = sheet_m.cell(row=last, column=3).value
span_start_last = sheet_m.cell(row=last - 1, column=2).value
span_end_last = sheet_m.cell(row=last - 1, column=3).value

span_s_str = str(span_start)[:-9].replace('-', '')  # spanを作成
span_e_str = str(span_end)[:-9].replace('-', '')
span = span_s_str + '-' + span_e_str
sheet_m.cell(row=last, column=1, value=span)
main_folder = 'H:/Youtube/' + span
os.makedirs(main_folder, exist_ok=True)  # フォルダ作成

span_time = datetime.timedelta(days=7)  # 次の週の期間を入力
next_span_e = span_end + span_time
sheet_m.cell(row=last + 1, column=2, value=span_end)
sheet_m.cell(row=last + 1, column=3, value=next_span_e)

sheet_m.cell(row=last, column=4, value=src)  # その他の情報を入力
exe_time = datetime.datetime.now()
sheet_m.cell(row=last, column=5, value=exe_time)
print('management done')


def change_length(x):
    element = str(x).split(':')
    result = int(element[0]) * 60 + int(element[1]) + int(element[2]) / 60
    result = round(result, 2)
    return result


# video_data作成
skiprows = list(range(1,6000))
df_base = pd.read_excel(src, sheet_name=1, index_col=0, skiprows=skiprows)
df_base['高評価数(再生数比x1000)'] = df_base['高評価数']/df_base['再生数'] * 1000
df_base['低評価数(再生数比x1000)'] = df_base['低評価数']/df_base['再生数'] * 1000
df_base['コメント数(再生数比x1000)'] = df_base['コメント数']/df_base['再生数'] * 1000
df_base['高評価-低評価比率'] = df_base['高評価数']/df_base['低評価数']
df_base['長さ(分)'] = df_base['長さ'].apply(change_length)
df_base['投稿日'] = df_base['投稿日時'].apply(lambda x: x[0:10].replace('/', '-'))
df_target = df_base[(str(span_end)[:10] > df_base['投稿日']) & (df_base['投稿日'] >= str(span_start)[:10])]
print('video_data done')

# 先週のvideo_data
df_target_last = df_base[(str(span_end_last)[:10] > df_base['投稿日']) & (df_base['投稿日'] >= str(span_start_last)[:10])]
print('video_data_last done')

# weekly作成
df_daily = pd.read_excel(src, sheet_name=2, index_col=0)  # 総再生数を計算
chousei = datetime.timedelta(hours=1)
chousei_2 = datetime.timedelta(days=1)
for i, col_name in enumerate(df_daily):
    col_date = datetime.datetime.strptime(col_name, '%Y/%m/%d %H:%M:%S')
    col_date = col_date - chousei
    print(col_date)
    if str(col_date)[:10] == str(span_start)[:10]:
        span_s_sum = df_daily[col_name].sum()
    if str(col_date + chousei_2)[:10] == str(span_end)[:10]:
        span_e_sum = df_daily[col_name].sum()
sheet_m.cell(row=last, column=17, value=(span_e_sum - span_s_sum))  # 総再生数

video_quantity = len(df_target)
sheet_m.cell(row=last, column=6, value=video_quantity)  # 動画数

sum_time = df_target['長さ(分)'].sum()  # 時間の合計
sheet_m.cell(row=last, column=7, value=sum_time)
sheet_m.cell(row=last, column=8, value=round(sum_time/video_quantity))  # 時間の平均

v_num = sheet_m.cell(row=last - 1, column=22).value  # 動画番号の更新
sheet_m.cell(row=last, column=22, value=v_num + 1)
i = 0
for col_name in df_target.iloc[:, 5:9]:
    sum_col = df_target[col_name].sum()
    sheet_m.cell(row=last, column=9 + i, value=sum_col)  # 記入
    i += 1
    sheet_m.cell(row=last, column=9 + i, value=round(sum_col / video_quantity, 2))  # 平均の記入
    i += 1

for i, col_name in enumerate(df_target.iloc[:, 11:15]): # 再生数比の各指標の記入(平均だけ)
    sum_col = df_target[col_name].sum()
    sheet_m.cell(row=last, column=18+i, value=round(sum_col / video_quantity, 2))

wb_management.save(management)
wb_management.save(main_folder + '/weekly.xlsx')
print('weekly done')

# 再生数の推移の部分
usecols = list(range(1, 1000))
df_time = pd.read_excel(src, sheet_name=3, usecols=usecols, index_col=0)


def make_vct(index, option):
    df_vc_trans = pd.DataFrame(index=index, columns=df_time.columns)
    for v_id in df_vc_trans.index:  # vc_transのvideoIDに一致するtimeの列を代入
        col = df_time[df_time.index == v_id].iloc[0, :]
        df_vc_trans.loc[v_id] = col
    df_vc_trans.to_excel(main_folder + '/vc_trans' + option + '.xlsx')
    print('vc_trans' + option + 'done')


make_vct(df_target.index, '')  # 今週の
make_vct(df_target_last.index, '_last')  # 今週の

# 伸び判定
df_time_rank = df_time[-300:].rank(numeric_only=True, ascending=False, method='min', na_option='top')
print(df_time_rank)
for v_id in df_target.index:
    print(v_id)
    df_target.loc[v_id, '伸び12時間'] = df_time_rank.loc[v_id, 12] - 1
    df_target.loc[v_id, '伸び48時間'] = df_time_rank.loc[v_id, 48] - 1
    df_target.loc[v_id, '伸び96時間'] = df_time_rank.loc[v_id, 96] - 1
df_target.to_excel(main_folder + '/video_data.xlsx')

# 先週の伸び判定
for v_id in df_target_last.index:
    print(v_id)
    df_target_last.loc[v_id, '伸び12時間'] = df_time_rank.loc[v_id, 12] - 1
    df_target_last.loc[v_id, '伸び48時間'] = df_time_rank.loc[v_id, 48] - 1
    df_target_last.loc[v_id, '伸び96時間'] = df_time_rank.loc[v_id, 96] - 1
    df_target_last.loc[v_id, '伸び144時間'] = df_time_rank.loc[v_id, 144] - 1
    df_target_last.loc[v_id, '伸び192時間'] = df_time_rank.loc[v_id, 192] - 1
    df_target_last.loc[v_id, '伸び240時間'] = df_time_rank.loc[v_id, 240] - 1
df_target_last.to_excel(main_folder + '/video_data_last.xlsx')
