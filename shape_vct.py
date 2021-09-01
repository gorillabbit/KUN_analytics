import glob

import openpyxl
import pandas as pd

folder_path = 'H:/Projects/basedata_KUN/vct/'
latest_vct = glob.glob(folder_path + '*_raw.xlsx')[-1]
df_latest_vct = pd.read_excel(latest_vct, index_col=0)

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(len(df_latest_vct)-1):
    print(i)
    shaped_row = df_latest_vct.iloc[i+1].dropna()

    for j, element in enumerate(shaped_row):
        sheet.cell(i+2, column=j+1, value=element)

for i in range(sheet.max_column-1):
    sheet.cell(1, column=2+i, value=i)  # 列と行の数字記入
sheet.cell(1, column=1, value='videoID')
wb.save(latest_vct[:-9]+ '_shaped.xlsx')