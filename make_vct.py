import datetime
import glob
import logging
import os
import shutil

import openpyxl
import pprint

import gspread as gs
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials as SACs

nowtime = str(datetime.datetime.now()).replace(':', '-').replace('.', '-')
logging.basicConfig(filename='H:/log/make_vct' + nowtime + '.log', level=logging.DEBUG)

api_scope = ['https://www.googleapis.com/auth/spreadsheets',  # 利用する API を指定する
             'https://www.googleapis.com/auth/drive']
credentials_path = 'H:/Youtube/analytics-315310-0789825df3f6.json'  # 先ほどダウンロードした json パスを指定する
os.path.join(os.path.expanduser('~'), 'path', 'to', 'analytics-315310-0789825df3f6.json')
credentials = SACs.from_json_keyfile_name(credentials_path, api_scope)  # json から Credentials 情報を取得
gspread_client = gs.authorize(credentials)  # 認可されたクライアントを得る

ss = gspread_client.open_by_key('1-8QnVNtgva-D10P6uBgbosStPUiwq82tzcdEiaiKx8U')
s_vct = ss.get_worksheet(2)
vct = s_vct.get_all_values()
hour_df = pd.DataFrame(vct)
logging.info(hour_df)
print(hour_df)

folder_path = 'H:/Projects/basedata_KUN/vct/'
os.makedirs(folder_path, exist_ok=True)  # vctフォルダー作成
nowtime = str(datetime.datetime.now()).replace(':', '-').replace('.', '-')
filepath = folder_path + 'vct_base' + nowtime + '_raw.xlsx'

latest_vct = glob.glob(folder_path + '*_raw.xlsx')[-1]

wb = openpyxl.load_workbook(latest_vct)
sheet = wb['1時間ごとの再生数']
for y,v_id in enumerate(hour_df[0]):
    print(y)
    sheet.cell(row=y+2, column=2, value=v_id)

col_num = sheet.max_column
for col_name in hour_df:
    if col_name == 0:
        continue
    print(col_name)
    col_num += 1
    for y, v_count in enumerate(hour_df[col_name]):
        sheet.cell(row=y+2, column=col_num, value=v_count)

for i in range(sheet.max_column):
    print(i)
    sheet.cell(row=1, column=i+2, value=str(i))

for i in range(sheet.max_row):
    sheet.cell(row=i+2, column=1, value=str(i))

wb.save(filepath)


s_vct.update_acell('B2', 'complete')
# s_vct.resize(cols=1)  # VideoID以外削除
logging.info(s_vct)
